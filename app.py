import streamlit as st
import pandas as pd
import math
import io
import re
import pdfplumber
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Set Page Config
st.set_page_config(page_title="Rekonsiliasi Meta & KBI", layout="wide")

st.title("📊 Aplikasi Rekonsiliasi Transaksi Meta vs KBI")
st.markdown("Unggah dokumen **List ACC**, **Trade Registry KBI (PDF)**, **Closed Trades (HTML)**, dan **Orders Report (HTML)** untuk melakukan pencocokan data secara otomatis.")

# --- SIDEBAR: FILE UPLOADER ---
with st.sidebar:
    st.header("📂 Upload Dokumen")
    file_acc = st.file_uploader("1. List ACC (.xlsx)", type=["xlsx"])
    file_kbi = st.file_uploader("2. Trade Registry KBI (.pdf)", type=["pdf"])
    file_closed = st.file_uploader("3. Closed Trades Report (.htm/.html)", type=["htm", "html"])
    file_orders = st.file_uploader("4. Orders Report (.htm/.html)", type=["htm", "html"])
    
    process_btn = st.button("🚀 Proses Rekonsiliasi", type="primary")

# --- CORE FUNCTIONS ---

def parse_kbi_pdf(pdf_file):
    pdf_trades = []
    current_type = 'Buy'
    
    with pdfplumber.open(pdf_file) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if not text: continue
            
            for line in text.split('\n'):
                line_lower = line.strip().lower()
                # Deteksi transisi section Buy/Sell
                if line_lower == 'buy':
                    current_type = 'Buy'
                elif line_lower == 'sell':
                    current_type = 'Sell'
                
                # Regex mendeteksi baris transaksi KBI
                # Pola: Jam (00:00:00) ... CC60XXXX ... Qty ... Kontrak (6 digit) ... Price
                match = re.search(r"(\d{2}:\d{2}:\d{2}).*?(CC\d+)\s+[A-Za-z0-9_]+\s+([\d\.]+)\s+\d{6}\s+([\d\,\.]+)", line.replace('|', ' '))
                if match:
                    time_kbi = match.group(1)
                    acc = match.group(2).replace('CC', '')
                    qty = float(match.group(3))
                    price = float(match.group(4).replace(',', ''))
                    
                    pdf_trades.append({
                        'Account': acc,
                        'Jam_KBI': time_kbi,
                        'Type_KBI': current_type,
                        'Price': price,
                        'Qty': qty,
                        'Matched': False
                    })
    return pdf_trades

def parse_meta_html(html_file, is_closed=True):
    soup = BeautifulSoup(html_file.read(), 'html.parser')
    rows = soup.find_all('tr', align="right")
    
    trades = []
    for r in rows:
        cols = r.find_all('td')
        if len(cols) > 7 and cols[0].text.strip().isdigit():
            deal = cols[0].text.strip()
            acc = cols[1].text.strip()
            
            if is_closed:
                # Kolom Closed Trades
                open_time = cols[3].text.strip()
                open_type = cols[4].text.strip().capitalize()
                vol = float(cols[6].text.strip())
                open_price = float(cols[7].text.strip().replace(' ', ''))
                
                close_time = cols[8].text.strip()
                close_price = float(cols[9].text.strip().replace(' ', ''))
                close_type = 'Sell' if open_type.lower() == 'buy' else 'Buy'
                
                if ' ' in open_time:
                    trades.append({'Deal': deal, 'Account': acc, 'Jam_Meta': open_time.split(' ')[1][:8], 
                                   'Type_Meta': open_type, 'Price': open_price, 'Qty': vol, 'Matched': False})
                if ' ' in close_time:
                    trades.append({'Deal': deal, 'Account': acc, 'Jam_Meta': close_time.split(' ')[1][:8], 
                                   'Type_Meta': close_type, 'Price': close_price, 'Qty': vol, 'Matched': False})
            else:
                # Kolom Orders Report
                time_val = cols[2].text.strip()
                typ = cols[3].text.strip().capitalize()
                vol = float(cols[5].text.strip())
                price = float(cols[6].text.strip().replace(' ', ''))
                jam = time_val.split(' ')[1][:8] if ' ' in time_val else 'Unknown'
                
                trades.append({'Deal': deal, 'Account': acc, 'Jam_Meta': jam, 'Type_Meta': typ, 'Price': price, 'Qty': vol})
    return trades

def build_excel(df):
    wb = Workbook()
    ws = wb.active
    ws.title = "Rekonsiliasi Final"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for r_idx, r in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(r, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = border
            if r_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                if isinstance(value, float) and 'Harga' not in df.columns[c_idx-1]:
                    cell.number_format = '0.00'
                cell.alignment = Alignment(horizontal="center")
                
                # Pewarnaan khusus untuk Catatan (Kolom ke-11)
                if c_idx == 11:
                    if "Seharusnya Posisi Masih Open" in str(value):
                        cell.font = Font(color="9C5700")
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    elif "Transaksi Salah tidak sesuai" in str(value):
                        cell.font = Font(color="9C0006")
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    elif "Posisi Seharusnya sudah Closed" in str(value):
                        cell.font = Font(color="006100")
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                
                if c_idx == 3 and value == 'TIDAK ADA':
                    cell.font = Font(color="9C0006")

    # Rapikan Lebar Kolom
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2

    excel_data = io.BytesIO()
    wb.save(excel_data)
    excel_data.seek(0)
    return excel_data

# --- MAIN EXECUTION ---
if process_btn:
    if not (file_acc and file_kbi and file_closed and file_orders):
        st.error("⚠️ Harap unggah KEMPAT file dokumen terlebih dahulu di menu sebelah kiri.")
    else:
        with st.spinner("Sedang memproses data..."):
            try:
                # 1. Map Account
                df_acc = pd.read_excel(file_acc)
                df_acc['Login'] = df_acc['Login'].astype(str).str.strip()
                df_acc['Nama'] = df_acc['Nama'].astype(str).str.strip()
                acc_map = dict(zip(df_acc['Login'], df_acc['Nama']))

                # 2. Extract Data
                pdf_trades = parse_kbi_pdf(file_kbi)
                meta_trades = parse_meta_html(file_closed, is_closed=True)
                order_trades = parse_meta_html(file_orders, is_closed=False)

                # 3. Rekonsiliasi Logic
                results = []
                for p_trade in pdf_trades:
                    acc = p_trade['Account']
                    price = p_trade['Price']
                    qty_kbi = p_trade['Qty']
                    type_kbi = p_trade['Type_KBI']
                    
                    best_match = None
                    # Cari yang match Harga & Qty
                    for m_trade in meta_trades:
                        if not m_trade['Matched'] and m_trade['Account'] == acc and m_trade['Type_Meta'] == type_kbi and math.isclose(m_trade['Price'], price, rel_tol=1e-5) and math.isclose(m_trade['Qty'], qty_kbi, rel_tol=1e-5):
                            best_match = m_trade
                            break
                            
                    # Cari yang match Harga Saja
                    if not best_match:
                        for m_trade in meta_trades:
                            if not m_trade['Matched'] and m_trade['Account'] == acc and m_trade['Type_Meta'] == type_kbi and math.isclose(m_trade['Price'], price, rel_tol=1e-5):
                                best_match = m_trade
                                break

                    if best_match:
                        best_match['Matched'] = True
                        deal_meta = best_match['Deal']
                        vol_meta = best_match['Qty']
                        type_meta = best_match['Type_Meta']
                        jam_meta = best_match['Jam_Meta']
                        catatan = 'Posisi Seharusnya sudah Closed'
                    else:
                        # Cek di Orders Report
                        found_in_orders = False
                        for o in order_trades:
                            if o['Account'] == acc and math.isclose(o['Price'], price, rel_tol=1e-5):
                                found_in_orders = True
                                deal_meta = o['Deal']
                                vol_meta = o['Qty']
                                type_meta = o['Type_Meta']
                                jam_meta = o['Jam_Meta']
                                catatan = 'Seharusnya Posisi Masih Open dan belum closed di Meta'
                                break
                                
                        if not found_in_orders:
                            deal_meta = 'TIDAK ADA'
                            vol_meta = 'TIDAK ADA'
                            type_meta = 'TIDAK ADA'
                            jam_meta = 'TIDAK ADA'
                            catatan = 'Transaksi Salah tidak sesuai'

                    results.append({
                        'Nomor Akun': acc,
                        'Nama Nasabah': acc_map.get(acc, '(Nama tidak ditemukan)'),
                        'Transaksi ID': deal_meta,
                        'Harga (Price)': price,
                        'Volume Meta (Lot)': vol_meta,
                        'Volume KBI (Lot)': qty_kbi,
                        'Jenis Transaksi Meta': type_meta,
                        'Jam Transaksi Meta': jam_meta,
                        'Jenis Transaksi KBI': type_kbi,
                        'Jam Transaksi KBI': p_trade['Jam_KBI'],
                        'Catatan': catatan
                    })

                df_final = pd.DataFrame(results)

                # Display Results
                st.success(f"✅ Berhasil memproses {len(df_final)} transaksi!")
                
                # Highlight in Streamlit UI
                def color_catatan(val):
                    color = 'green' if 'Closed' in val else ('orange' if 'Open' in val else 'red')
                    return f'color: {color}; font-weight: bold;'
                
                st.dataframe(df_final.style.map(color_catatan, subset=['Catatan']), use_container_width=True)

                # Download Button
                excel_buffer = build_excel(df_final)
                st.download_button(
                    label="📥 Download Hasil Rekonsiliasi (Excel)",
                    data=excel_buffer,
                    file_name="Rekonsiliasi_Final_Meta_KBI.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary"
                )

            except Exception as e:
                st.error(f"❌ Terjadi kesalahan saat memproses data: {e}")